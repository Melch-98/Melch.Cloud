from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Brand Colors ──
GOLD = 'C8B89A'
DARK_BG = '0A0A0A'
CARD_BG = '1A1A1A'
HEADER_BG = '111111'
WHITE = 'F5F5F8'
MUTED = '999999'
GREEN = '10B981'
RED = 'EF4444'
BLUE = '5DADE2'

hfont = Font(name='Arial', bold=True, size=11, color=WHITE)
hfill = PatternFill('solid', fgColor=HEADER_BG)
gold_font = Font(name='Arial', bold=True, size=11, color=GOLD)
gold_fill = PatternFill('solid', fgColor='1C1A16')
body_font = Font(name='Arial', size=10, color=WHITE)
muted_font = Font(name='Arial', size=10, color=MUTED)
green_font = Font(name='Arial', bold=True, size=10, color=GREEN)
money_font = Font(name='Arial', size=10, color=WHITE)
blue_font = Font(name='Arial', bold=True, size=10, color=BLUE)
dark_fill = PatternFill('solid', fgColor=DARK_BG)
card_fill = PatternFill('solid', fgColor=CARD_BG)
alt_fill = PatternFill('solid', fgColor='141414')
thin_border = Border(
    bottom=Side(style='thin', color='333333')
)
gold_border = Border(
    bottom=Side(style='medium', color=GOLD)
)

def style_range(ws, row, cols, font=body_font, fill=dark_fill, border=thin_border, align=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        if border: cell.border = border
        if align: cell.alignment = align

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════════════════════════════
# SHEET 1: Infrastructure Cost Breakdown
# ═══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Infrastructure Costs'
ws.sheet_properties.tabColor = GOLD

set_col_widths(ws, [32, 18, 16, 16, 16, 40])

# Dark background for all visible cells
for r in range(1, 60):
    style_range(ws, r, 6, fill=dark_fill, border=None)

# Title
ws['A1'] = 'MELCH.CLOUD — INFRASTRUCTURE COST BREAKDOWN'
ws['A1'].font = Font(name='Arial', bold=True, size=14, color=GOLD)
ws['A1'].fill = dark_fill
ws['A2'] = 'April 2026'
ws['A2'].font = muted_font
ws['A2'].fill = dark_fill

# Section 1: Current Stack
r = 4
headers = ['Service', 'Current Tier', 'Monthly Cost', 'Annual Cost', 'Pro Tier Cost', 'Notes']
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=r, column=i, value=h)
    cell.font = hfont
    cell.fill = PatternFill('solid', fgColor='1C1A16')
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='left')

# Data rows
services = [
    ['Vercel (Hosting)', 'Hobby (Free)', 0, '=C5*12', 20, 'Free: 10s fn timeout, 100GB BW. Pro: 60s fn, 1TB BW, $20 usage credit'],
    ['Supabase (DB + Auth)', 'Free', 0, '=C6*12', 25, 'Free: 500MB DB, 50k MAU, pauses after 7d inactive. Pro: 8GB, no pause'],
    ['Vercel Blob (CDN)', 'Included', 0, '=C7*12', 0, '$0.023/GB storage + $0.05/GB transfer. Minimal at current volume'],
    ['Resend (Email)', 'Free', 0, '=C8*12', 0, 'Free: 3,000 emails/mo (100/day). More than enough for notifications'],
    ['BetterStack (Uptime)', 'Free', 0, '=C9*12', 0, 'Free: 10 monitors, 3-min checks, status page. Currently sufficient'],
    ['Anthropic API (Claude)', 'Pay-as-you-go', 5, '=C10*12', 5, 'Sonnet 4: $3/$15 per 1M tokens. ~$5/mo estimate for Ad Lab usage'],
    ['Meta Marketing API', 'Free', 0, '=C11*12', 0, 'No API cost — included with Meta Business account'],
    ['Klaviyo API', 'Free', 0, '=C12*12', 0, 'API access included with client Klaviyo accounts'],
    ['Domain (melch.cloud)', 'Annual', 1.50, '=C13*12', 1.50, '~$18/yr via Porkbun (.cloud TLD)'],
]

for i, row_data in enumerate(services):
    r = 5 + i
    fill = card_fill if i % 2 == 0 else alt_fill
    for j, val in enumerate(row_data):
        cell = ws.cell(row=r, column=j+1, value=val)
        cell.font = body_font
        cell.fill = fill
        cell.border = thin_border
        if j in [2, 3, 4]:
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal='right')
        if j == 5:
            cell.font = muted_font

# Totals row
r = 14
ws.cell(row=r, column=1, value='TOTAL MONTHLY').font = gold_font
ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor='1C1A16')
ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor='1C1A16')
ws.cell(row=r, column=3, value='=SUM(C5:C13)').font = Font(name='Arial', bold=True, size=11, color=GREEN)
ws.cell(row=r, column=3).fill = PatternFill('solid', fgColor='1C1A16')
ws.cell(row=r, column=3).number_format = '$#,##0.00'
ws.cell(row=r, column=3).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=4, value='=SUM(D5:D13)').font = Font(name='Arial', bold=True, size=11, color=GREEN)
ws.cell(row=r, column=4).fill = PatternFill('solid', fgColor='1C1A16')
ws.cell(row=r, column=4).number_format = '$#,##0.00'
ws.cell(row=r, column=4).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=5, value='=SUM(E5:E13)').font = Font(name='Arial', bold=True, size=11, color=BLUE)
ws.cell(row=r, column=5).fill = PatternFill('solid', fgColor='1C1A16')
ws.cell(row=r, column=5).number_format = '$#,##0.00'
ws.cell(row=r, column=5).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=6, value='Pro upgrade adds 60s function timeout + no DB pause').font = muted_font
ws.cell(row=r, column=6).fill = PatternFill('solid', fgColor='1C1A16')
for c in range(1, 7):
    ws.cell(row=r, column=c).border = gold_border

# Annual totals
r = 15
ws.cell(row=r, column=1, value='TOTAL ANNUAL').font = gold_font
ws.cell(row=r, column=1).fill = dark_fill
ws.cell(row=r, column=3, value='=C14*12').font = Font(name='Arial', bold=True, size=11, color=GREEN)
ws.cell(row=r, column=3).fill = dark_fill
ws.cell(row=r, column=3).number_format = '$#,##0.00'
ws.cell(row=r, column=3).alignment = Alignment(horizontal='right')
ws.cell(row=r, column=5, value='=E14*12').font = Font(name='Arial', bold=True, size=11, color=BLUE)
ws.cell(row=r, column=5).fill = dark_fill
ws.cell(row=r, column=5).number_format = '$#,##0.00'
ws.cell(row=r, column=5).alignment = Alignment(horizontal='right')

# ── Section 2: What You Get ──
r = 18
ws.cell(row=r, column=1, value='FULL FEATURE INVENTORY').font = Font(name='Arial', bold=True, size=13, color=GOLD)
ws.cell(row=r, column=1).fill = dark_fill

r = 19
for i, h in enumerate(['Feature', 'Category', 'Status', 'Powered By'], 1):
    cell = ws.cell(row=r, column=i, value=h)
    cell.font = hfont
    cell.fill = PatternFill('solid', fgColor='1C1A16')
    cell.border = gold_border

features = [
    ['Meta Ads analytics dashboard', 'Analytics', 'Live', 'Meta API + Supabase'],
    ['Ad Perspective (outlier/percentile analysis)', 'Analytics', 'Live', 'Meta API'],
    ['Copy Analysis (headline/body performance)', 'Analytics', 'Live', 'Meta API'],
    ['Creative detail sidecar (video + image)', 'Analytics', 'Live', 'Meta API + Vercel Blob'],
    ['Full-res thumbnails via CDN', 'Analytics', 'Live', 'Vercel Blob'],
    ['CSV export (Ad Perspective + Copy Analysis)', 'Analytics', 'Live', 'Client-side'],
    ['Ad Lab (AI copy generation)', 'Creative', 'Live', 'Anthropic Claude'],
    ['Creative upload pipeline', 'Operations', 'Live', 'Supabase'],
    ['Submission tracking + batch management', 'Operations', 'Live', 'Supabase'],
    ['Klaviyo email/SMS campaign calendar', 'Calendar', 'Live', 'Klaviyo API'],
    ['Team management + role-based access', 'Admin', 'Live', 'Supabase Auth'],
    ['Strategist permissions (upload/download/delete)', 'Admin', 'Live', 'Supabase'],
    ['Feature request board + voting', 'Community', 'Live', 'Supabase'],
    ['Changelog / release notes page', 'Community', 'Live', 'Static'],
    ['Admin email notifications on submissions', 'Notifications', 'Live', 'Resend'],
    ['Uptime monitoring (3-min checks, US+EU)', 'Infrastructure', 'Live', 'BetterStack'],
    ['Public status page (status.melch.cloud)', 'Infrastructure', 'Live', 'BetterStack'],
    ['Friendly error messages + data freshness', 'DX', 'Live', 'Client-side'],
    ['Meta token health check (admin)', 'Admin', 'Live', 'Meta API'],
]

for i, feat in enumerate(features):
    r = 20 + i
    fill = card_fill if i % 2 == 0 else alt_fill
    for j, val in enumerate(feat):
        cell = ws.cell(row=r, column=j+1, value=val)
        cell.font = body_font if j != 2 else green_font
        cell.fill = fill
        cell.border = thin_border

# ═══════════════════════════════════════════════════════════════
# SHEET 2: Service Packaging
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Service Packaging')
ws2.sheet_properties.tabColor = GOLD

set_col_widths(ws2, [36, 20, 20, 20, 40])

for r in range(1, 50):
    style_range(ws2, r, 5, fill=dark_fill, border=None)

ws2['A1'] = 'MELCH.CLOUD — SERVICE PACKAGING OPTIONS'
ws2['A1'].font = Font(name='Arial', bold=True, size=14, color=GOLD)
ws2['A2'] = 'How to bundle with your agency services'
ws2['A2'].font = muted_font

# Tier headers
r = 4
tiers = ['', 'Starter', 'Growth', 'Premium']
for i, t in enumerate(tiers):
    cell = ws2.cell(row=r, column=i+1, value=t)
    cell.font = Font(name='Arial', bold=True, size=12, color=GOLD if i > 0 else WHITE)
    cell.fill = PatternFill('solid', fgColor='1C1A16')
    cell.border = gold_border
    cell.alignment = Alignment(horizontal='center')
ws2.cell(row=r, column=1).alignment = Alignment(horizontal='left')
ws2.cell(row=r, column=5, value='Notes').font = hfont
ws2.cell(row=r, column=5).fill = PatternFill('solid', fgColor='1C1A16')
ws2.cell(row=r, column=5).border = gold_border

# Pricing section
rows = [
    ['Monthly add-on price (suggested)', 250, 500, 1000, 'Price to client on top of retainer'],
    ['Your infrastructure cost', '=\'Infrastructure Costs\'!C14', '=\'Infrastructure Costs\'!C14', '=\'Infrastructure Costs\'!E14', 'Current free-tier or Pro-tier cost'],
    ['Your margin', '=B5-B6', '=C5-C6', '=D5-D6', ''],
    ['Margin %', '=B7/B5', '=C7/C5', '=D7/D5', '98%+ margin on Starter/Growth, 95%+ on Premium'],
    ['', '', '', '', ''],
    ['INCLUDED FEATURES', '', '', '', ''],
    ['Meta Ads dashboard + analytics', 'Yes', 'Yes', 'Yes', 'Core value prop'],
    ['Ad Perspective analysis', 'Yes', 'Yes', 'Yes', 'Percentile + outlier breakdown'],
    ['Copy Analysis', 'No', 'Yes', 'Yes', 'Upsell from Starter to Growth'],
    ['Creative sidecar (video + image)', 'Yes', 'Yes', 'Yes', 'Full-res via CDN'],
    ['CSV exports', 'No', 'Yes', 'Yes', 'Data portability'],
    ['Ad Lab (AI copy generation)', 'No', 'No', 'Yes', 'Premium-only differentiator'],
    ['Klaviyo calendar integration', 'No', 'Yes', 'Yes', 'Email/SMS campaign visibility'],
    ['Creative upload pipeline', 'No', 'No', 'Yes', 'Full ops workflow'],
    ['Feature request access', 'Yes', 'Yes', 'Yes', 'Community engagement'],
    ['Status page access', 'Yes', 'Yes', 'Yes', 'Transparency / trust'],
    ['Dedicated Slack alerts', 'No', 'No', 'Yes', 'Premium support signal'],
    ['Custom branding (white-label)', 'No', 'No', 'Yes', 'Future upsell opportunity'],
    ['', '', '', '', ''],
    ['REVENUE PROJECTIONS', '', '', '', ''],
    ['Clients on this tier', 3, 5, 2, 'Example allocation across 10 clients'],
    ['Monthly revenue per tier', '=B5*B25', '=C5*C25', '=D5*D25', ''],
    ['Total monthly platform revenue', '', '', '=B26+C26+D26', ''],
    ['Total annual platform revenue', '', '', '=D27*12', ''],
    ['Annual infrastructure cost', '', '', '=\'Infrastructure Costs\'!C15', ''],
    ['Annual net profit from platform', '', '', '=D28-D29', ''],
]

for i, row_data in enumerate(rows):
    r = 5 + i
    fill = card_fill if i % 2 == 0 else alt_fill
    
    # Section headers
    if row_data[0] in ['INCLUDED FEATURES', 'REVENUE PROJECTIONS']:
        fill = PatternFill('solid', fgColor='1C1A16')
        for j, val in enumerate(row_data):
            cell = ws2.cell(row=r, column=j+1, value=val)
            cell.font = gold_font
            cell.fill = fill
            cell.border = gold_border
        continue
    
    # Empty rows
    if row_data[0] == '':
        for j in range(5):
            ws2.cell(row=r, column=j+1).fill = dark_fill
        continue
    
    for j, val in enumerate(row_data):
        cell = ws2.cell(row=r, column=j+1, value=val)
        cell.fill = fill
        cell.border = thin_border
        
        if j == 0:
            cell.font = body_font
        elif j == 4:
            cell.font = muted_font
        elif val == 'Yes':
            cell.font = green_font
            cell.alignment = Alignment(horizontal='center')
        elif val == 'No':
            cell.font = Font(name='Arial', size=10, color='555555')
            cell.alignment = Alignment(horizontal='center')
        elif isinstance(val, str) and val.startswith('='):
            cell.font = money_font
            cell.alignment = Alignment(horizontal='right')
        elif isinstance(val, (int, float)):
            cell.font = money_font
            cell.alignment = Alignment(horizontal='right')

# Format currency cells
for r_idx in [5, 6, 7, 26, 27, 28, 29, 30]:
    for c in range(2, 5):
        cell = ws2.cell(row=r_idx, column=c)
        cell.number_format = '$#,##0.00'

# Format percentage
for c in range(2, 5):
    ws2.cell(row=8, column=c).number_format = '0.0%'

# Highlight the totals
for r_idx in [27, 28, 30]:
    cell = ws2.cell(row=r_idx, column=4)
    cell.font = Font(name='Arial', bold=True, size=11, color=GREEN)

ws2.cell(row=30, column=1).font = gold_font

out = '/sessions/quirky-relaxed-archimedes/mnt/Melch.Cloud/melch-cloud-cost-breakdown.xlsx'
wb.save(out)
print(f'Saved to {out}')
